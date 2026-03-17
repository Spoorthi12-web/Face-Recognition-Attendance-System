import pandas as pd
import datetime
import os
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import DB_CONFIG, REPORTS_DIR

def get_connection():
    return mysql.connector.connect(**DB_CONFIG)

def generate_excel_report(date=None):
    if date is None:
        date = datetime.date.today()

    print(f"Generating Excel report for: {date}")

    conn   = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT p.full_name, p.department, p.email,
               a.time_in, a.time_out, a.status,
               TIMEDIFF(a.time_out, a.time_in) as hours_worked
        FROM attendance_log a
        JOIN persons p ON p.person_id = a.person_id
        WHERE a.date = %s
        ORDER BY a.time_in
    """, (date,))
    present_rows = cursor.fetchall()

    cursor.execute("""
        SELECT p.full_name, p.department, p.email
        FROM persons p
        WHERE p.person_id NOT IN (
            SELECT person_id FROM attendance_log
            WHERE date = %s
        )
    """, (date,))
    absent_rows = cursor.fetchall()

    cursor.execute("""
        SELECT a.date, COUNT(*) as present_count
        FROM attendance_log a
        WHERE a.date >= DATE_SUB(%s, INTERVAL 7 DAY)
        GROUP BY a.date
        ORDER BY a.date
    """, (date,))
    weekly_rows = cursor.fetchall()

    cursor.close()
    conn.close()

    total_present = len(present_rows)
    total_absent  = len(absent_rows)
    total         = total_present + total_absent
    pct           = round((total_present / total * 100), 1) if total > 0 else 0

    os.makedirs(REPORTS_DIR, exist_ok=True)
    filename = f"Attendance_Report_{date}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Present", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Absent",  index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Weekly",  index=False)

    wb = load_workbook(filepath)

    green_fill   = PatternFill("solid", fgColor="1B5E20")
    red_fill     = PatternFill("solid", fgColor="B71C1C")
    blue_fill    = PatternFill("solid", fgColor="0D47A1")
    orange_fill  = PatternFill("solid", fgColor="E65100")
    present_fill = PatternFill("solid", fgColor="E8F5E9")
    absent_fill  = PatternFill("solid", fgColor="FFEBEE")
    header_fill  = PatternFill("solid", fgColor="E3F2FD")
    white_fill   = PatternFill("solid", fgColor="FFFFFF")
    lgray_fill   = PatternFill("solid", fgColor="FAFAFA")

    white_bold  = Font(color="FFFFFF", bold=True, size=12)
    title_font  = Font(bold=True, size=16, color="0D47A1")
    normal_font = Font(size=11)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    thin = Border(
        left   = Side(style="thin",   color="BDBDBD"),
        right  = Side(style="thin",   color="BDBDBD"),
        top    = Side(style="thin",   color="BDBDBD"),
        bottom = Side(style="thin",   color="BDBDBD")
    )
    thick = Border(
        left   = Side(style="medium", color="0D47A1"),
        right  = Side(style="medium", color="0D47A1"),
        top    = Side(style="medium", color="0D47A1"),
        bottom = Side(style="medium", color="0D47A1")
    )

    # ── SHEET 1: SUMMARY ─────────────────────────────────────
    ws1 = wb["Summary"]
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("B2:G2")
    ws1["B2"]           = "ATTENDANCE REPORT"
    ws1["B2"].font      = title_font
    ws1["B2"].alignment = center

    ws1.merge_cells("B3:G3")
    ws1["B3"]           = f"Date: {date}  |  Generated: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws1["B3"].font      = Font(size=11, color="546E7A")
    ws1["B3"].alignment = center

    cards = [
        ("B5", "C6", "Total Staff",  total,         blue_fill),
        ("D5", "E6", "Present",      total_present, green_fill),
        ("F5", "G6", "Absent",       total_absent,  red_fill),
        ("H5", "I6", "Attendance %", f"{pct}%",     orange_fill),
    ]
    for start, end, label, value, fill in cards:
        ws1.merge_cells(f"{start}:{end}")
        ws1[start]           = f"{label}\n{value}"
        ws1[start].fill      = fill
        ws1[start].font      = Font(color="FFFFFF", bold=True, size=14)
        ws1[start].alignment = center
        ws1[start].border    = thick

    headers1 = ["Category", "Count", "Percentage"]
    for col, h in enumerate(headers1, start=2):
        cell            = ws1.cell(row=8, column=col, value=h)
        cell.fill       = blue_fill
        cell.font       = white_bold
        cell.alignment  = center
        cell.border     = thin

    summary_data = [
        ("Total Registered", total,         "100%"),
        ("Present Today",    total_present, f"{pct}%"),
        ("Absent Today",     total_absent,  f"{round(100-pct,1)}%"),
    ]
    for i, (cat, cnt, pct_val) in enumerate(summary_data, start=9):
        fill = present_fill if "Present" in cat else (absent_fill if "Absent" in cat else lgray_fill)
        for col, val in zip([2, 3, 4], [cat, cnt, pct_val]):
            cell            = ws1.cell(row=i, column=col, value=val)
            cell.fill       = fill
            cell.alignment  = center
            cell.border     = thin
            cell.font       = normal_font

    ws1.column_dimensions["A"].width = 3
    for col in ["B","C","D","E","F","G","H","I"]:
        ws1.column_dimensions[col].width = 14
    ws1.row_dimensions[5].height = 50
    ws1.row_dimensions[6].height = 50

    # ── SHEET 2: PRESENT ─────────────────────────────────────
    ws2 = wb["Present"]
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    ws2["A1"]           = f"PRESENT EMPLOYEES — {date}"
    ws2["A1"].font      = Font(bold=True, size=14, color="1B5E20")
    ws2["A1"].fill      = present_fill
    ws2["A1"].alignment = center
    ws2["A1"].border    = thin
    ws2.row_dimensions[1].height = 30

    headers2 = ["#", "Name", "Department", "Email", "Punch In", "Punch Out", "Hours Worked"]
    for col, h in enumerate(headers2, start=1):
        cell            = ws2.cell(row=2, column=col, value=h)
        cell.fill       = green_fill
        cell.font       = white_bold
        cell.alignment  = center
        cell.border     = thin

    if present_rows:
        for row_idx, row in enumerate(present_rows, start=3):
            fill   = present_fill if row_idx % 2 == 0 else white_fill
            values = [
                row_idx - 2,
                row[0],
                row[1],
                row[2],
                str(row[3]) if row[3] else "—",
                str(row[4]) if row[4] else "—",
                str(row[6]) if row[6] else "—"
            ]
            for col, val in enumerate(values, start=1):
                cell            = ws2.cell(row=row_idx, column=col, value=val)
                cell.fill       = fill
                cell.alignment  = center
                cell.border     = thin
                cell.font       = normal_font
    else:
        ws2.merge_cells("A3:G3")
        ws2["A3"]           = "No attendance records for today!"
        ws2["A3"].font      = Font(bold=True, color="B71C1C", size=12)
        ws2["A3"].alignment = center
        ws2["A3"].fill      = absent_fill

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 25
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 14

    # ── SHEET 3: ABSENT ──────────────────────────────────────
    ws3 = wb["Absent"]
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:D1")
    ws3["A1"]           = f"ABSENT EMPLOYEES — {date}"
    ws3["A1"].font      = Font(bold=True, size=14, color="B71C1C")
    ws3["A1"].fill      = absent_fill
    ws3["A1"].alignment = center
    ws3["A1"].border    = thin
    ws3.row_dimensions[1].height = 30

    headers3 = ["#", "Name", "Department", "Email"]
    for col, h in enumerate(headers3, start=1):
        cell            = ws3.cell(row=2, column=col, value=h)
        cell.fill       = red_fill
        cell.font       = white_bold
        cell.alignment  = center
        cell.border     = thin

    if absent_rows:
        for row_idx, row in enumerate(absent_rows, start=3):
            fill   = absent_fill if row_idx % 2 == 0 else white_fill
            values = [row_idx - 2, row[0], row[1], row[2]]
            for col, val in enumerate(values, start=1):
                cell            = ws3.cell(row=row_idx, column=col, value=val)
                cell.fill       = fill
                cell.alignment  = center
                cell.border     = thin
                cell.font       = normal_font
    else:
        ws3.merge_cells("A3:D3")
        ws3["A3"]           = "No absences today — full attendance!"
        ws3["A3"].font      = Font(bold=True, color="2E7D32", size=12)
        ws3["A3"].alignment = center
        ws3["A3"].fill      = present_fill

    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 25

    # ── SHEET 4: WEEKLY ──────────────────────────────────────
    ws4 = wb["Weekly"]
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:C1")
    ws4["A1"]           = "WEEKLY ATTENDANCE SUMMARY"
    ws4["A1"].font      = Font(bold=True, size=14, color="0D47A1")
    ws4["A1"].fill      = header_fill
    ws4["A1"].alignment = center
    ws4["A1"].border    = thin
    ws4.row_dimensions[1].height = 30

    headers4 = ["Date", "Present Count", "Day"]
    for col, h in enumerate(headers4, start=1):
        cell            = ws4.cell(row=2, column=col, value=h)
        cell.fill       = blue_fill
        cell.font       = white_bold
        cell.alignment  = center
        cell.border     = thin

    if weekly_rows:
        for row_idx, row in enumerate(weekly_rows, start=3):
            day_name = row[0].strftime("%A") if hasattr(row[0], "strftime") else str(row[0])
            fill     = header_fill if row_idx % 2 == 0 else white_fill
            values   = [str(row[0]), row[1], day_name]
            for col, val in enumerate(values, start=1):
                cell            = ws4.cell(row=row_idx, column=col, value=val)
                cell.fill       = fill
                cell.alignment  = center
                cell.border     = thin
                cell.font       = normal_font
    else:
        ws4.merge_cells("A3:C3")
        ws4["A3"]           = "No weekly data available yet!"
        ws4["A3"].font      = Font(bold=True, color="546E7A", size=12)
        ws4["A3"].alignment = center
        ws4["A3"].fill      = lgray_fill

    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 14

    # ── SAVE ─────────────────────────────────────────────────
    wb.save(filepath)

    print(f"")
    print(f"Report generated successfully!")
    print(f"----------------------------------")
    print(f"Date        : {date}")
    print(f"Present     : {total_present}")
    print(f"Absent      : {total_absent}")
    print(f"Total       : {total}")
    print(f"Attendance  : {pct}%")
    print(f"----------------------------------")
    print(f"File saved  : {filepath}")
    print(f"Open reports/ folder to view!")
    return filepath

if __name__ == "__main__":
    generate_excel_report()
