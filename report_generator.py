# report_generator.py — Generate Excel Attendance Report

import pandas as pd
import datetime
import os
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from config import DB_CONFIG, REPORTS_DIR

def get_connection():
    return mysql.connector.connect(**DB_CONFIG)

def generate_report(date=None):
    # Use today's date if not specified
    if date is None:
        date = datetime.date.today()

    print(f"Generating report for: {date}")

    conn   = get_connection()
    cursor = conn.cursor()

    # ── Get Present List ──────────────────────────────────────────
    cursor.execute("""
        SELECT p.full_name, p.department, 
               a.time_in, a.time_out, a.status
        FROM attendance_log a
        JOIN persons p ON p.person_id = a.person_id
        WHERE a.date = %s
        ORDER BY a.time_in
    """, (date,))
    present_rows = cursor.fetchall()

    # ── Get Absent List ───────────────────────────────────────────
    cursor.execute("""
        SELECT full_name, department
        FROM persons
        WHERE person_id NOT IN (
            SELECT person_id FROM attendance_log
            WHERE date = %s
        )
    """, (date,))
    absent_rows = cursor.fetchall()

    cursor.close()
    conn.close()

    # ── Build DataFrames ──────────────────────────────────────────
    df_present = pd.DataFrame(present_rows,
        columns=["Name", "Department", "Punch In", "Punch Out", "Status"])

    df_absent = pd.DataFrame(absent_rows,
        columns=["Name", "Department"])
    df_absent["Status"] = "ABSENT"

    # ── Summary ───────────────────────────────────────────────────
    total_present = len(df_present)
    total_absent  = len(df_absent)
    total         = total_present + total_absent

    print(f"Present : {total_present}")
    print(f"Absent  : {total_absent}")
    print(f"Total   : {total}")

    # ── Save to Excel ─────────────────────────────────────────────
    os.makedirs(REPORTS_DIR, exist_ok=True)
    filename = f"attendance_{date}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

        # ── Summary Sheet ─────────────────────────────────────────
        df_summary = pd.DataFrame({
            "Category" : ["Total Students", "Present", "Absent", "Date"],
            "Value"    : [total, total_present, total_absent, str(date)]
        })
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

        # ── Present Sheet ─────────────────────────────────────────
        df_present.to_excel(writer, sheet_name="Present", index=False)

        # ── Absent Sheet ──────────────────────────────────────────
        df_absent.to_excel(writer, sheet_name="Absent", index=False)

    # ── Apply Excel Styling ───────────────────────────────────────
    wb = load_workbook(filepath)

    # Colors
    green_fill  = PatternFill("solid", fgColor="00C853")
    red_fill    = PatternFill("solid", fgColor="FF1744")
    blue_fill   = PatternFill("solid", fgColor="1565C0")
    white_font  = Font(color="FFFFFF", bold=True, size=12)
    header_font = Font(bold=True, size=11)
    center      = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left   = Side(style="thin"),
        right  = Side(style="thin"),
        top    = Side(style="thin"),
        bottom = Side(style="thin")
    )

    # Style Summary Sheet
    ws = wb["Summary"]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    for row in ws.iter_rows():
        for cell in row:
            cell.border    = thin_border
            cell.alignment = center
    for cell in ws[1]:
        cell.fill = blue_fill
        cell.font = white_font

    # Style Present Sheet
    ws = wb["Present"]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    for row in ws.iter_rows():
        for cell in row:
            cell.border    = thin_border
            cell.alignment = center
    for cell in ws[1]:
        cell.fill = green_fill
        cell.font = white_font

    # Style Absent Sheet
    ws = wb["Absent"]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    for row in ws.iter_rows():
        for cell in row:
            cell.border    = thin_border
            cell.alignment = center
    for cell in ws[1]:
        cell.fill = red_fill
        cell.font = white_font

    wb.save(filepath)

    print(f"---------------------------")
    print(f"Report saved: {filepath}")
    print(f"Open the reports/ folder to see it!")
    return filepath

# Run directly
if __name__ == "__main__":
    generate_report()

